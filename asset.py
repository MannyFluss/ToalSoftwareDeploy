import csv;
import re;
import copy;
import pandas as pd;
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font



csv_pathname : str = "./examples/21024.csv"
excelsheet_pathname : str = "./examples/21024 Monitor for MannyCopy.xlsx"
excelsheet_result_pathname : str = "./examples/21024 Monitor for MannyCopyResult.xlsx"

#the csv file format can be very different, plus it may only have partial informationm
#of the current excel sheet.
csvRowBlocks : list = []

csvWorkingRowBlock : list = []

currentWorkingNEELValues : dict = {}

#return a list of the current working block
def readCSV(filePathName,_startingLine): 
    with open(filePathName,'r') as file:
        currBlock : list = []
        currBlockCount : int = 0
        pattern = re.compile(r'^MP\d*\.?\d*')
        reader = csv.reader(file) 
        # for row in reader:
        #     print(row)
        #     #empty row
        #     if row == []:
        #         continue
        #     #check if mp
        #     if pattern.match(row[0]) or pattern.match(row[4]):
        #         rowCopy = row.copy()
        #         currBlock.append(rowCopy)
        #         currBlockCount += 1
        #     else:
        #         #this means this is the end of a current scan
        #         if (currBlockCount > 0):
        #             blockCopy = copy.deepcopy(currBlock)
        #             csvRowBlocks.append(blockCopy)
        #             #clear the working blocks
        #             currBlockCount = 0
        #             currBlock.clear()
        # blockCopy = copy.deepcopy(currBlock)
        # csvRowBlocks.append(blockCopy)
        # #print(csvRowBlocks[len(csvRowBlocks)-1])
        # csvWorkingRowBlock = copy.deepcopy(csvRowBlocks[len(csvRowBlocks)-1])
        # #print(csvWorkingRowBlock)
        #now i will just do the same thing but only MP's after a certain number
        readMonitorsFlag = False
        currBlock.clear()
        currBlockCount = 0
        #print(_startingLine)
        for row in reader:
            
            if row == []:
                continue
            #print(row)
            if str(row[0]) == _startingLine:
                readMonitorsFlag = True
                print("hope you see this message")
            if readMonitorsFlag == False:
                continue
            #these are the mps that we will actually use
            if pattern.match(row[0]) or pattern.match(row[4]):
                rowCopy = row.copy()
                currBlock.append(rowCopy)

        if readMonitorsFlag == False:
            raise Exception("CSV File failed to read any Input")
            
        return copy.deepcopy(currBlock)



    #able to put into their individual blocks, find dates next.
def printCSVRowBlocks():
    for block in csvRowBlocks:
        for row in block:
            print(row)
        print()

def alignMonitorPoints(_CSVpoints:list, _workingExcelColumn:list) -> list:
    CSVPoints = copy.deepcopy(_CSVpoints)

    CSVPoints = [sublist[1:4] for sublist in CSVPoints]
    to_return = [None] * len(_workingExcelColumn)
    #print(to_return)

    #print(CSVPoints)

    workingExcelColumn = copy.deepcopy(_workingExcelColumn)


    #csvSubsetLists = []
    workingExcelColumnSubsetLists = []
    #splitting sets into subsets
    for i in range(0, len(workingExcelColumn), 3):
        workingExcelColumnSubsetLists.append(workingExcelColumn[i:i+3])

    
  #  print(CSVPoints)  
  #  print(workingExcelColumnSubsetLists)
    for csvTriplet in CSVPoints:
        smallestDelta = 1000
        currentSmallestSlot=0
        smallestTriplet = ["error","error","error"]
        #find where the current csvTriplet will be placed
        for i,excelTriplet in enumerate(workingExcelColumnSubsetLists):

            workingDelta = abs(float(csvTriplet[0])-float(excelTriplet[0]))
            workingDelta = abs(float(csvTriplet[1])-float(excelTriplet[1]))
            workingDelta = abs(float(csvTriplet[2])-float(excelTriplet[2]))

            if (workingDelta < smallestDelta):
                smallestDelta = workingDelta
                currentSmallestSlot = i
                smallestTriplet = excelTriplet
        if (smallestDelta > 5):
            print("delta exceeds allowed amount, new point to be added")
            to_return.append(csvTriplet[0])
            to_return.append(csvTriplet[1])
            to_return.append(csvTriplet[2])
        else:
            to_return[currentSmallestSlot*3]=csvTriplet[0]
            to_return[currentSmallestSlot*3+1]=csvTriplet[1]
            to_return[currentSmallestSlot*3+2]=csvTriplet[2]
    for i, item in enumerate(to_return):
        if item==None:
            to_return[i] = "NA"
#    for item in to_return:
 #       print(item)

    
    return to_return

def getFirstWorkingColumn(_ws , _monitorPointCount:int) -> list:
    #go through each column starting at E, check if there is a number
    to_return = [None] * (_monitorPointCount * 3)
    for i in range(0,_monitorPointCount):
        curr_row = 3+(i*3)
        curr_column = 5 # start at col 5, work to right
        j = 0
        while(True):
            j+=1
            if (j > 100000):
                raise Exception("monitor point unable to be located at target cell curr row: " +str(curr_row) )
            #check if current column works
            North = _ws.cell(row=curr_row,column=curr_column)
            East = _ws.cell(row=curr_row+1,column=curr_column)
            Elevation = _ws.cell(row=curr_row+2,column=curr_column)
            if (North.value == None or East.value == None or Elevation.value == None):
                curr_column +=1
                continue                
            if (North.value == "N" or East.value == "E" or Elevation.value == "EL"):
                curr_column +=1
                continue
            to_return[i*3] = North.value
            to_return[i*3 + 1] = East.value
            to_return[i*3 + 2] = Elevation.value
            break
    return to_return
def getMostRecentColumnValues(ws, targetCell, monitorPointCount) ->list:
    mostRecentColumnValues = []
    pointCount = (monitorPointCount) * 3
    for i in range(pointCount):
        #iterate through row starting at target cell until you hit a value
        targetRow = 3+i # doesnt change
        currentColumn = targetCell.column # changes
        deltaFlagCount = 0 # 0 means, DEEL, 1 means Delta number, 2 is what we want
        while(True):
            currentCell = ws.cell(row=targetRow,column=currentColumn)
            currentColumn -= 1
            if deltaFlagCount == 0 or deltaFlagCount == 1:
                deltaFlagCount = deltaFlagCount + 1
                continue
            #we are on what can possibly be a value we want
            pattern = re.compile(r"\d+(\.\d+)?")
            match = pattern.search(str(currentCell.value))
            if match:
                mostRecentColumnValues.append(str(currentCell.value))
                deltaFlagCount = 0 
                break
            deltaFlagCount = 0

    return mostRecentColumnValues

defaultStyle=Side(border_style='medium',color='000000')

def applySquareBorder(ws, topLeftCorner,bottomRightCorner,sideStyle = defaultStyle):
    #construct border styles
    cell_range = ws[str(topLeftCorner.coordinate)+":"+str(bottomRightCorner.coordinate)]
    mostLeft = topLeftCorner.column
    mostTop = topLeftCorner.row
    mostBottom = bottomRightCorner.row
    mostRight = bottomRightCorner.column
    
    for row in cell_range:
        for cell in row:
            #check if its on each border
            _top,_left,_right,_bottom = None,None,None,None
            if cell.row == mostBottom:
                _bottom = sideStyle
            if cell.column == mostLeft:
                _left = sideStyle
            if cell.column == mostRight:
                _right = sideStyle
            if cell.row == mostTop:
                _top = sideStyle
            cell.border = Border(right=_right,left=_left,bottom=_bottom,top=_top)
   

def proccessExcelSheet(csvFilePath:str,excelFilePath:str, monitorDate = "The Date", startingLine = "10000"):
    print("checkpoint1")
    csvPoints:list = readCSV(csvFilePath,startingLine)
    print("checkpoint2")

    wb = load_workbook(excelFilePath)
    ws = wb['MONCTRL'] #this may pose problems if not all workbooks are monctrl
    column = ws['A']
    column_list = [cell.value for cell in column]
    
    #get the current amount of monitor points/ rows we are operating on.
    monitorPointCount : int = 0
    for element in column_list:
        if type(element) == int:
            monitorPointCount = element
    #find the target column
    # Insert a new column for the copied data
    targetCell = None
    for cell in ws[1]:
        if cell.value == "Overall":
            targetCell = ws.cell(row=cell.row,column=cell.column-1)
            break
    #setting up the values to be inserted
    entryDate = monitorDate
    FirstColumnValues = [None,None] 
    SecondColumnValues = [entryDate,"Location"]
    ThirdColumnValues = [None,"Delta"]
    for i in range(0,monitorPointCount+1):
        FirstColumnValues = FirstColumnValues + ["N","E","EL"]
    
    print("checkpoint2")

    column_letter = targetCell.column_letter
    ws.insert_cols(targetCell.column,3)
    targetCell = ws.cell(targetCell.row,targetCell.column-3)
    secondTargetCell = ws.cell(targetCell.row,targetCell.column+1)
    thirdTargetCell = ws.cell(targetCell.row,targetCell.column+2)

    OverallDeltaCellFeet = ws.cell(row=3,column=targetCell.column+5)
    OverallDeltaCellInches = ws.cell(row=3,column=targetCell.column+8)
    
    OverallDeltaCellFeetColumn = [None] * ((i+1)*3)
    OverallDeltaCellInchesColumn = [None] * ((i+1)*3)
    
    
    for i in range(0,i):
        currFeet = ws.cell(row=OverallDeltaCellFeet.row + i, column=OverallDeltaCellFeet.column)
        currInches = ws.cell(row=OverallDeltaCellInches.row + i, column=OverallDeltaCellInches.column)
        OverallDeltaCellFeetColumn[i] = currFeet.value
        OverallDeltaCellInchesColumn[i] = currInches.value
    

######
    #format cells
    #for some reason refuses to work unless I do this
    mergedCell1 = secondTargetCell.column_letter + str(2)
    mergedCell2 = thirdTargetCell.column_letter + str(2)    
    ws.merge_cells(mergedCell1+":"+mergedCell2)
    ws.unmerge_cells(mergedCell1+":"+mergedCell2)
    mostRecentColumnValues = getMostRecentColumnValues(ws,targetCell,monitorPointCount+1)

    newValues = alignMonitorPoints(csvPoints,mostRecentColumnValues)
    SecondColumnValues = SecondColumnValues + newValues

    for i, item in enumerate(newValues):
        if (newValues[i]=="NA"):
            ThirdColumnValues.append("NA")
        else:
            ThirdColumnValues.append(round(float(newValues[i])-float(mostRecentColumnValues[i]),3))    


    for i, value in enumerate(FirstColumnValues):
        cell = ws[column_letter + str(i+1)]
        cell.value = value

    
    for i, value in enumerate(SecondColumnValues):
        cell = ws[secondTargetCell.column_letter+str(i+1)]
        cell.value = value
    for i, value in enumerate(ThirdColumnValues):
        cell = ws[thirdTargetCell.column_letter+str(i+1)]
        cell.value = value
    #we need to get the correct values in the overall Delta
    firstWorkingColumn = getFirstWorkingColumn(ws,monitorPointCount+1)
    postInsertTarget = ws.cell(row=targetCell.row,column=targetCell.column + 3)
    latestWorkingColumn = getMostRecentColumnValues(ws,postInsertTarget,monitorPointCount+1)
    newDeltasFeet = [round((float(latestWorkingColumn[i])-float(firstWorkingColumn[i])),3) for i in range(0,len(latestWorkingColumn))]
    newDeltasinches = [round((float(latestWorkingColumn[i])-float(firstWorkingColumn[i])) * 12,2) for i in range(0,len(latestWorkingColumn))]

    for i in range(0,len(latestWorkingColumn)):
        deltaFeetCurr = ws.cell(row=3+i,column=OverallDeltaCellFeet.column)
        deltaInchesCurr = ws.cell(row=3+i,column=OverallDeltaCellInches.column)

        deltaFeetCurr.value = newDeltasFeet[i]
        deltaInchesCurr.value = newDeltasinches[i]


    #we need to format the values
    column_list = [(ws.cell(row=targetCell.row+3,column=targetCell.column+i).column_letter) for i in range(0,9)]


    ws.column_dimensions[column_list[0]].width = 2.43
    ws.column_dimensions[column_list[1]].width = 9
    ws.column_dimensions[column_list[2]].width = 6
    ws.column_dimensions[column_list[3]].width = 1.86

    ws.column_dimensions[column_list[4]].width = 2.57
    ws.column_dimensions[column_list[5]].width = 9.29
    ws.column_dimensions[column_list[6]].width = 1.86
    ws.column_dimensions[column_list[7]].width = 2.57
    ws.column_dimensions[column_list[8]].width = 9.29

    merge1 = column_list[7]+"1"
    merge2 = column_list[8]+"1"
    ws.merge_cells(merge1+":"+merge2)

    merge1 = column_list[7]+"2"
    merge2 = column_list[8]+"2"
    ws.merge_cells(merge1+":"+merge2)
    
    
    applySquareBorder(ws,targetCell,secondTargetCell)
    #apply all the borders
    bottomRightBorderCell = ws.cell(column=targetCell.column+2,row=2)
    applySquareBorder(ws,targetCell,bottomRightBorderCell)
    cellBorderCount = max(13,monitorPointCount+1)
    for i in range(0,cellBorderCount):
        curr_cell = ws.cell(row=3+(3*i),column=targetCell.column)
        other_cell = ws.cell(row=3+2+(3*i),column=targetCell.column+2)
        applySquareBorder(ws,curr_cell,other_cell)
        curr_cell = ws.cell(row=3+(3*i),column=targetCell.column+2)
        other_cell = ws.cell(row=3+2+(3*i),column=targetCell.column+2)
        applySquareBorder(ws,curr_cell,other_cell)
    
    cell_align = Alignment(horizontal='center')
    ws.cell(row=1,column=targetCell.column+1).alignment = cell_align
    ws.cell(row=2,column=targetCell.column+1).alignment = cell_align
    ws.cell(row=2,column=targetCell.column+2).alignment = cell_align

    cell_align = Alignment(horizontal='right')
    topLeftWorkingCell = ws.cell(row=3,column=targetCell.column+1).coordinate
    bottomRightWorkingCell = ws.cell(row=3+2+(cellBorderCount*3),column=targetCell.column+3).coordinate
    for rows in ws[topLeftWorkingCell+":"+bottomRightWorkingCell]:
        for cell in rows:
            cell.number_format = '0.000'
            cell.alignment = cell_align
    wb.save(excelFilePath[:-5] +"results.xlsx")



    pass

def main():
    # Your code here
    #readCSV(,"TR503.16")
    #printCSVRowBlocks()
    proccessExcelSheet("C:/Users/manny/Desktop/Git Folder/ToalSoftware/examples/21024.csv",
                       excelFilePath="C:/Users/manny/Desktop/Git Folder/ToalSoftware/examples/21024 Monitor for Manny.xlsx",
                        monitorDate="2",startingLine="TR503.16")
    pass
if __name__ == '__main__':
    main()

