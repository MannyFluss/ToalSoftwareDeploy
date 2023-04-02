import csv;
import re;
import math;
import copy;
import pandas as pd;
import openpyxl
from openpyxl import load_workbook
#from openpyxl.formula import parser
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

class CellTriplet:
    # North :float = 0.0
    # East :float = 0.0
    # Elevation :float = 0.0
    def __str__(self):
        return str(self.North) +" "+str(self.East)+" "+str(self.Elevation)
    def __init__(self,_northing=0.0,_easting=0.0,_elevation=0.0 ):
        n = _northing
        e = _easting
        el = _elevation
        # try:
        #     float(_northing)
        #     float(_easting)
        #     float(_elevation)
        # finally:
        #     formula_string_n = n
        #     formula_string_e = e
        #     formula_string_el = el

        #     tokensn = parser.parse(formula_string_n)
        #     tokense = parser.parse(formula_string_e)
        #     tokensel = parser.parse(formula_string_el)

        #     n = tokensn.evaluate()
        #     e = tokense.evaluate()
        #     el = tokensel.evaluate()
        #     pass

        self.North = float(n)
        self.East = float(e)
        self.Elevation = float(el)

    def getNorth(self):
        return self.North
    
    @staticmethod
    def deltaBetweenCells(_targ1 : "CellTriplet",_targ2 :"CellTriplet" ) -> float:
        to_return = 0

        
        #print(_targ2.North)
        to_return += abs(_targ1.North - _targ2.North)
        to_return += abs(_targ1.East - _targ2.East)
        to_return += abs(_targ1.Elevation - _targ2.Elevation) 
        return to_return
    @staticmethod
    def duplicate(_target : "CellTriplet")->"CellTriplet":
        
        to_return = CellTriplet(_target.North,_target.East,_target.Elevation)
        return to_return

class CellColumn:
    def __init__(self):
        self.CellList =  []
        return
    def CellListAppend(self,_toAdd : CellTriplet):
        self.CellList.append(_toAdd)
    def getCellCount(self)->int:
        return len(self.CellList)
    def clearCellList(self):
        self.CellList.clear()
    def __iter__(self):
        self.currentIndex = 0
        #enum for 012
        self.cellState = 0
        return self
    def getIndex(self,_index) -> float:
        neel = _index%3
        cellIndex = int(math.floor(_index/3))
        if cellIndex >= len(self.CellList):
            return 0.0
        if neel ==0:
            return self.CellList[cellIndex].North
        if neel == 1:
            
            return self.CellList[cellIndex].East
        if neel == 2:
            return self.CellList[cellIndex].Elevation 

    def __next__(self):
        if self.currentIndex == len(self.CellList):
            raise StopIteration
        currCell : CellTriplet = self.CellList[self.currentIndex]
        if self.cellState == 0:
            self.cellState = 1
            if (currCell):
                return currCell.North
            return None
        if self.cellState == 1:
            self.cellState = 2
            if(currCell):
                return currCell.East
            return None
        if self.cellState == 2:
            self.cellState = 0
            self.currentIndex += 1
            if (currCell):
                return currCell.Elevation
            return None
#these should be constant variables, advanced options that should not ever need to be changed, but may need to be in a special case
theDate = "0/0/0"
default_cell_count = 13
default_print_bottom = 57
determinantTargetRow = 1
determinantTargetWord = "Overall"
startingRow = 3
startingColumn = 5
smallestDeltaAllowed = 3
#csv data
csvUnorderedResults : CellColumn = CellColumn()
csvOrderedValues : CellColumn = CellColumn()
#excelsheetData
cellCount : int = 0
targetColumn : int = 0
excelMostRecentWorkingColumn : CellColumn = CellColumn()
excelFirstWorkingColumn : CellColumn = CellColumn()



def readCSV(fileName:str, startingLine:str, endingLine:str):
    with open(fileName,'r') as file:
        pattern = re.compile(r'^MP\d*\.?\d*')
        reader = csv.reader(file) 
        
        readMonitorPointsFlag = False
        for row in reader:
            
            if row == []:
                continue
            if row[0] == startingLine:
                readMonitorPointsFlag = True
            if readMonitorPointsFlag==False:
                continue
            if pattern.match(row[0]) or pattern.match(row[4]):
                
                toAdd = CellTriplet(row[1],row[2],row[3])
                csvUnorderedResults.CellListAppend(toAdd)
            if str(row[0]) == endingLine:
                break
        if readMonitorPointsFlag==False:
            raise Exception("error no monitorpoints were read")

def getMostRecentWorkingValues(_ws):     
    ws = _ws
    excelMostRecentWorkingColumn.clearCellList()
    for i in range(0,cellCount):
        currentTargetRow = startingRow + i*3
        currentTargetColumn = targetColumn - 2
        while(True):
            currentCell = ws.cell(row=currentTargetRow,column=currentTargetColumn)
            pattern = re.compile(r"\d+(\.\d+)?")
            match = pattern.search(str(currentCell.value))
            if (match):
                val1 = currentCell.value
                val2 = ws.cell(row=currentTargetRow+1,column=currentTargetColumn).value
                val3 = ws.cell(row=currentTargetRow+2,column=currentTargetColumn).value
                toAdd = CellTriplet(val1,val2,val3)
                excelMostRecentWorkingColumn.CellListAppend(toAdd)
                break
            currentTargetColumn -= 3

def getFirstWorkingValeus(_ws):
    
    ws = _ws
    excelFirstWorkingColumn.clearCellList()
    for i in range(0,cellCount):
        currentTargetRow = startingRow + i*3
        currentTargetColumn = startingColumn
        while(True):
            currentCell = ws.cell(row=currentTargetRow,column=currentTargetColumn)    
            pattern = re.compile(r"\d+(\.\d+)?")
            match = pattern.search(str(currentCell.value))
            if (match and currentCell.value > 1):
                val1 = currentCell.value
                val2 = ws.cell(row=currentTargetRow+1,column=currentTargetColumn).value
                val3 = ws.cell(row=currentTargetRow+2,column=currentTargetColumn).value
                toAdd = CellTriplet(val1,val2,val3)
                excelFirstWorkingColumn.CellListAppend(toAdd)
                break
            currentTargetColumn += 1

def readExcelSheet(filePath:str,worksheetName:str="MONCTRL"):
    #find the target column
    global cellCount
    global targetColumn

    wb = load_workbook(filePath, data_only=True)
    ws = wb[worksheetName]
    for cell in ws[determinantTargetRow]:
        if cell.value == determinantTargetWord:
            targetColumn = cell.column - 1
            break
    #find the amount of cells that we are currently working with
    currentCell = ws.cell(startingRow,targetColumn-1)
    tempCellCount = 0
    while(True):
        if currentCell.value == None:
            break
        tempCellCount += 1
        currentCell = ws.cell(currentCell.row+1,targetColumn-1)
    cellCount = int(tempCellCount/3)
    getMostRecentWorkingValues(ws)
    getFirstWorkingValeus(ws)

def orderCSVValues():
    global cellCount
    orderedList = [None] * cellCount
    for currentCSVValue in csvUnorderedResults.CellList:
        currentSmallestCSVValue : CellTriplet = None
        currentSmallestDelta = 10000000.0
        current_i = 0
        for i, excelValue in enumerate(excelMostRecentWorkingColumn.CellList):
            currCSV : CellTriplet = currentCSVValue
            currExcel : CellTriplet = excelValue
            #delta = CellTriplet.deltaBetweenCells(excelValue,currentCSVValue) 
            delta=0
            delta += abs(float(currCSV.North) - float(currExcel.North))
            delta += abs(float(currCSV.East) - float(currExcel.East))
            delta += abs(float(currCSV.Elevation) - float(currExcel.Elevation))

            if delta < currentSmallestDelta:
               currentSmallestCSVValue = currentCSVValue
               currentSmallestDelta = delta
               current_i = i

        if currentSmallestDelta > smallestDeltaAllowed:
            print("delta exceeds smallest allowed amount")
            cellCount += 1
            orderedList.append(CellTriplet(currentSmallestCSVValue.North,currentSmallestCSVValue.East,currentSmallestCSVValue.Elevation))
        else:
            #print()
            #print(currentSmallestDelta)
            #print("delta doesnt exceeds smallest allowed amount")
            orderedList[current_i] = CellTriplet(currentSmallestCSVValue.North,currentSmallestCSVValue.East,currentSmallestCSVValue.Elevation)

    for cell in orderedList:
        csvOrderedValues.CellListAppend(cell)

    pass

def insertRowsAndWrite(filePath:str,worksheetName:str="MONCTRL"):
    wb = load_workbook(filePath)
    ws = wb[worksheetName]
    ws.insert_cols(targetColumn,3)

    newValueTarget = ws.cell(startingRow,targetColumn+1)
    for i,value in enumerate(csvOrderedValues):
        to_insert = "NA"
        if value == None or value == "":
            to_insert = "NA"
        else:
            to_insert = float(value)
        newValueTarget.value = to_insert
        leftOfValue = ws.cell(row=newValueTarget.row,column=newValueTarget.column-1)
        if i%3==0:
            leftOfValue.value = "N"
        if i%3==1:
            leftOfValue.value = "E"
        if i%3==2:
            leftOfValue.value = "EL"
        rightOfValue = ws.cell(row=newValueTarget.row,column=newValueTarget.column+1)
        if to_insert=="NA" or excelMostRecentWorkingColumn.getIndex(i) == 0:
            rightOfValue.value = "NA"
        else:
            rightOfValue.value = float(to_insert) - excelMostRecentWorkingColumn.getIndex(i)
        
        newValueTarget = ws.cell(row=newValueTarget.row+1,column=newValueTarget.column)
        

    wb.save(filePath[:-5] +"results.xlsx")
    
def updateOverallDeltaValues(filePath:str,worksheetName:str="MONCTRL"):
    wb = load_workbook(filePath)
    ws = wb[worksheetName]
    #there is a possiblity this is larger than the working values of before(new point added)
    getMostRecentWorkingValues(ws)
    getFirstWorkingValeus(ws)
    currCellFeet = ws.cell(row=startingRow,column=targetColumn+5)
    currCellInches = ws.cell(row=startingRow,column=targetColumn+8)
    for i,recentValue in enumerate(excelMostRecentWorkingColumn):
        firstValue = excelFirstWorkingColumn.getIndex(i)
        valueToInsertFeet = ""
        valueToInsertInches = ""
        if firstValue == 0.0:
            valueToInsertFeet = "NA"
            valueToInsertInches = "NA"
        else:
            valueToInsertFeet = float(round(float(recentValue) - float(firstValue),3))
            valueToInsertInches = float(round((float(recentValue) - float(firstValue))*12,2))
        currCellFeet.value = valueToInsertFeet
        currCellInches.value = valueToInsertInches
        currCellFeet = ws.cell(row=currCellFeet.row+1,column=currCellFeet.column)
        currCellInches = ws.cell(row=currCellInches.row+1,column=currCellInches.column)

    #for now just apply style here too
    applyStyle(ws)

    wb.save(filePath[:-5] +"results.xlsx")
    pass
# apply the right amount of nums to each value
# apply the square borders to each of the things
# apply alignment to the text
date_row = 1
location_delta_row = 2
def applyStyle(_ws):
    ws = _ws
    #apply proper width
    column_list = [(ws.cell(row=startingRow,column=targetColumn+i).column_letter) for i in range(0,9)]
    ws.column_dimensions[column_list[0]].width = 2.43
    ws.column_dimensions[column_list[1]].width = 11
    ws.column_dimensions[column_list[2]].width = 6
    ws.column_dimensions[column_list[3]].width = 1.86

    ws.column_dimensions[column_list[4]].width = 2.57
    ws.column_dimensions[column_list[5]].width = 9.29
    ws.column_dimensions[column_list[6]].width = 1.86
    ws.column_dimensions[column_list[7]].width = 2.57
    ws.column_dimensions[column_list[8]].width = 9.29
    #cell mergings
    merge1 = column_list[7]+str(date_row)
    merge2 = column_list[8]+str(date_row)
    ws.merge_cells(merge1+":"+merge2)
    merge1 = column_list[7]+str(location_delta_row)
    merge2 = column_list[8]+str(location_delta_row)
    ws.merge_cells(merge1+":"+merge2)

    #
    mergedCell1 = ws.cell(row=startingRow,column=targetColumn+1).column_letter + str(2)
    mergedCell2 = ws.cell(row=startingRow,column=targetColumn+2).column_letter + str(2)
    ws.merge_cells(mergedCell1+":"+mergedCell2)
    ws.unmerge_cells(mergedCell1+":"+mergedCell2)
    #fill in misc vals
    dateCell = ws.cell(row=date_row,column=targetColumn+1)
    dateCell.value = theDate
    
    locationCell = ws.cell(row=location_delta_row,column=targetColumn+1)
    deltaCell = ws.cell(row=location_delta_row,column=targetColumn+2)

    locationCell.value = "Location"
    deltaCell.value = "Delta"

    cell_align = Alignment(horizontal='center')
    dateCell.alignment = cell_align
    locationCell.alignment = cell_align
    deltaCell.alignment = cell_align
    #apply number format, and right allignment
    cell_align = Alignment(horizontal='right')
    topLeftWorkingCell = ws.cell(row=startingRow,column=targetColumn)
    bottomRightWorkingCell = ws.cell(row=startingRow+2+(cellCount*3),column=targetColumn+5)
    for rows in ws[topLeftWorkingCell.coordinate+":"+bottomRightWorkingCell.coordinate]:
        for cell in rows:
            cell.number_format = "0.000"
            cell.alignment = cell_align
    #apply the borders
    #update the gui
    topLeftCell = ws.cell(row=date_row,column=targetColumn)
    bottomRightCell = ws.cell(row=location_delta_row,column=targetColumn+2)
    applySquareBorder(ws,topLeftCell,bottomRightCell)
    for i in range(0,max(cellCount,default_cell_count)):
        
        topLeftCell = ws.cell(row=startingRow+(i*3),column=targetColumn)
        bottomRightCell = ws.cell(row=startingRow+(i*3)+2,column=targetColumn+2)
        applySquareBorder(ws,topLeftCell,bottomRightCell)
        topLeftCell = ws.cell(row=startingRow+(i*3),column=targetColumn+2)
        bottomRightCell = ws.cell(row=startingRow+2+(i*3),column=targetColumn+2)
        applySquareBorder(ws,topLeftCell,bottomRightCell)
        
        
    
    topLeftCell = ws.cell(row=1,column=1)
    bottomRightCell = ws.cell(row=default_print_bottom,column=targetColumn+8)
    ws.print_area = topLeftCell.coordinate+":"+bottomRightCell.coordinate

defaultStyle=Side(border_style='medium',color='000000')

def applySquareBorder(ws, topLeftCorner,bottomRightCorner,sideStyle = defaultStyle):
    #construct border styles
    cell_range = ws[str(topLeftCorner.coordinate)+":"+str(bottomRightCorner.coordinate)]
    mostLeft = topLeftCorner.column
    mostTop = topLeftCorner.row
    mostBottom = bottomRightCorner.row
    mostRight = bottomRightCorner.column
    for row in cell_range:
        for cell in row:
            #check if its on each border
            _top,_left,_right,_bottom = None,None,None,None
            if cell.row == mostBottom:
                _bottom = sideStyle
            if cell.column == mostLeft:
                _left = sideStyle
            if cell.column == mostRight:
                _right = sideStyle
            if cell.row == mostTop:
                _top = sideStyle
            cell.border = Border(right=_right,left=_left,bottom=_bottom,top=_top)
   


def main():
    
    testCase = "C:/Users/manny/Desktop/Git Folder/ToalSoftwareDeploy/ToalSoftwareDeploy/examples/21024 Monitor for MannyCopy.xlsx"
    testCase1 = "C:/Users/manny/Desktop/Git Folder/ToalSoftwareDeploy/ToalSoftwareDeploy/examples/21024.csv"
    testCase2 = "C:/Users/manny/Desktop/Git Folder/ToalSoftwareDeploy/ToalSoftwareDeploy/examples/21024 Monitor for MannyCopyresult.xlsx"

    test2 = "C:/Users/manny/Desktop/Git Folder/ToalSoftwareDeploy/ToalSoftwareDeploy/examples/14814  Monitor.xlsx"
    test1 = "C:/Users/manny/Desktop/Git Folder/ToalSoftwareDeploy/ToalSoftwareDeploy/examples/14814.csv"

    
    readCSV(testCase1,"TR503.16","")
    
    readExcelSheet(testCase, "MONCTRL")
    orderCSVValues()
    insertRowsAndWrite(testCase,"MONCTRL")
    updateOverallDeltaValues(testCase[:-5] +"results.xlsx","MONCTRL")

    # # print()
    # for i ,element in enumerate(csvOrderedValues):
    #      print(" "+ str(excelFirstWorkingColumn.getIndex(i)))
    # for element in excelFirstWorkingColumn:
    #     print(element)



    # for element in excelMostRecentWorkingColumn:
    #     print(element)

    # for element in csvUnorderedResults:
    #     print(element)

if __name__ == '__main__':
    main()